from django.contrib import admin
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import path
from django import forms
from .models import Department, JobTitle, Employee, SalaryStructure, Organization
import openpyxl
import pandas as pd
import io
from datetime import datetime
from decimal import Decimal
import re
import chardet


def parse_number_value(value):
    """
    Parse various number formats including quoted values and comma separators.
    Handles: "7,000", "7000", 7000, 7.00, "7.00", etc.
    """
    if pd.isna(value) or value == '' or str(value).strip() == '':
        return None

    # Convert to string and strip whitespace
    str_value = str(value).strip()

    # Remove quotes if present
    if str_value.startswith('"') and str_value.endswith('"'):
        str_value = str_value[1:-1]
    elif str_value.startswith("'") and str_value.endswith("'"):
        str_value = str_value[1:-1]

    # Handle scientific notation (like 1.10056E+12)
    if 'E+' in str_value.upper() or 'E-' in str_value.upper():
        try:
            return float(str_value)
        except ValueError:
            pass

    # Remove commas used as thousand separators
    # But preserve decimal commas (European format)
    if ',' in str_value and '.' in str_value:
        # Both comma and dot present - assume comma is thousand separator
        str_value = str_value.replace(',', '')
    elif ',' in str_value:
        # Only comma present - could be thousand separator or decimal
        comma_parts = str_value.split(',')
        if len(comma_parts) == 2 and len(comma_parts[1]) <= 2:
            # Likely decimal comma (e.g., "7,50")
            str_value = str_value.replace(',', '.')
        else:
            # Likely thousand separator (e.g., "7,000" or "1,000,000")
            str_value = str_value.replace(',', '')

    try:
        # Try to convert to float first, then to appropriate type
        float_value = float(str_value)

        # If it's a whole number, return as int
        if float_value.is_integer():
            return int(float_value)
        else:
            return float_value
    except ValueError:
        # If all else fails, return the original string
        return str_value


def detect_file_encoding(file):
    """Detect file encoding using chardet"""
    try:
        file.seek(0)
        raw_data = file.read()
        file.seek(0)

        # Use chardet to detect encoding
        detection = chardet.detect(raw_data)
        encoding = detection.get('encoding', 'utf-8')
        confidence = detection.get('confidence', 0)

        print(f"Detected encoding: {encoding} (confidence: {confidence:.2f})")

        # If confidence is low, fall back to common encodings
        if confidence < 0.7:
            print(f"Low confidence ({confidence:.2f}), will try multiple encodings")
            return None

        return encoding
    except Exception as e:
        print(f"Error detecting encoding: {e}")
        return None


def clean_string_value(value):
    """Clean string values by removing quotes and extra whitespace"""
    if pd.isna(value) or value == '':
        return None

    str_value = str(value).strip()

    # Remove quotes if present
    if str_value.startswith('"') and str_value.endswith('"'):
        str_value = str_value[1:-1]
    elif str_value.startswith("'") and str_value.endswith("'"):
        str_value = str_value[1:-1]

    # Return None for empty strings after cleaning
    if str_value.strip() == '' or str_value.lower() in ['nan', 'null', 'none']:
        return None

    return str_value.strip()


class BulkImportForm(forms.Form):
    """Form for bulk importing employees"""
    file = forms.FileField(
        label="Upload File",
        help_text="Upload Excel (.xlsx, .xls) or CSV file with employee data",
        widget=forms.FileInput(attrs={
            'accept': '.xlsx,.xls,.csv',
            'class': 'form-control'
        })
    )


@admin.register(Organization)
class OrganizationAdmin(admin.ModelAdmin):
    list_display = ['name', 'organization_type', 'short_name', 'ministry', 'sector', 'city', 'is_active']
    list_filter = ['organization_type', 'is_active', 'country', 'city', 'sector']
    search_fields = ['name', 'short_name', 'trading_name', 'kra_pin', 'registration_number', 'ministry']
    fieldsets = (
        ('Basic Information', {
            'fields': ('organization_type', 'name', 'short_name', 'trading_name', 'is_active')
        }),
        ('Organizational Structure', {
            'fields': ('ministry', 'sector'),
            'description': 'For government entities, specify the parent ministry and sector'
        }),
        ('Contact Information', {
            'fields': ('address_line_1', 'address_line_2', 'city', 'postal_code', 'country', 'phone_number', 'email', 'website')
        }),
        ('Registration & Compliance', {
            'fields': ('kra_pin', 'registration_number', 'registration_authority', 'nssf_employer_number', 'shif_employer_number')
        }),
        ('Payroll Settings', {
            'fields': ('default_pay_day', 'logo')
        }),
    )


@admin.register(Department)
class DepartmentAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'organization', 'organization_type', 'created_at']
    list_filter = ['organization', 'organization__organization_type', 'created_at']
    search_fields = ['name', 'code', 'organization__name', 'organization__short_name']
    ordering = ['organization__name', 'name']

    def organization_type(self, obj):
        return obj.organization.get_organization_type_display()
    organization_type.short_description = 'Organization Type'


@admin.register(JobTitle)
class JobTitleAdmin(admin.ModelAdmin):
    list_display = ['title', 'created_at']
    list_filter = ['created_at']
    search_fields = ['title']
    ordering = ['title']


class SalaryStructureInline(admin.StackedInline):
    model = SalaryStructure
    extra = 0


@admin.register(Employee)
class EmployeeAdmin(admin.ModelAdmin):
    list_display = [
        'payroll_number', 'full_name', 'department', 'job_title',
        'employment_type', 'basic_salary', 'date_hired', 'is_active'
    ]
    list_filter = [
        'department', 'job_title', 'employment_type', 'is_active',
        'gender', 'marital_status', 'date_hired'
    ]
    search_fields = [
        'payroll_number', 'first_name', 'last_name', 'email',
        'national_id', 'kra_pin', 'nssf_number', 'shif_number'
    ]
    ordering = ['payroll_number']

    # Exclude payroll_number from forms since it's auto-generated
    exclude = ['payroll_number']

    fieldsets = (
        ('Personal Information', {
            'fields': (
                'first_name', 'middle_name', 'last_name',
                'date_of_birth', 'gender', 'marital_status', 'national_id'
            )
        }),
        ('Contact Information', {
            'fields': ('email', 'phone_number', 'address')
        }),
        ('Employment Information', {
            'fields': (
                'department', 'job_title', 'employment_type',
                'date_hired', 'date_terminated', 'is_active', 'basic_salary'
            )
        }),
        ('Statutory Information', {
            'fields': ('kra_pin', 'nssf_number', 'shif_number')
        }),
        ('Banking Information', {
            'fields': ('bank_name', 'bank_branch', 'account_number')
        }),
        ('System Information', {
            'fields': ('user',),
            'classes': ('collapse',)
        })
    )

    def get_readonly_fields(self, request, obj=None):
        """Return readonly fields - payroll_number only for existing employees"""
        if obj:  # Editing existing employee - show payroll_number as readonly
            return ['payroll_number']
        return []  # No readonly fields when adding new employee

    def get_exclude(self, request, obj=None):
        """Exclude payroll_number from add form, but show it in edit form"""
        if not obj:  # Adding new employee
            return ['payroll_number']
        return []  # Don't exclude anything when editing existing employee

    def get_fieldsets(self, request, obj=None):
        """Return different fieldsets for add vs change forms"""
        if obj:  # Editing existing employee - include payroll_number
            return (
                ('Personal Information', {
                    'fields': (
                        'payroll_number', 'first_name', 'middle_name', 'last_name',
                        'date_of_birth', 'gender', 'marital_status', 'national_id'
                    )
                }),
                ('Contact Information', {
                    'fields': ('email', 'phone_number', 'address')
                }),
                ('Employment Information', {
                    'fields': (
                        'department', 'job_title', 'employment_type',
                        'date_hired', 'date_terminated', 'is_active', 'basic_salary'
                    )
                }),
                ('Statutory Information', {
                    'fields': ('kra_pin', 'nssf_number', 'shif_number')
                }),
                ('Banking Information', {
                    'fields': ('bank_name', 'bank_branch', 'account_number')
                }),
                ('System Information', {
                    'fields': ('user',),
                    'classes': ('collapse',)
                })
            )
        return super().get_fieldsets(request, obj)  # Use default fieldsets for add form

    inlines = [SalaryStructureInline]
    actions = ['bulk_import_employees']

    def full_name(self, obj):
        return obj.full_name
    full_name.short_description = 'Full Name'

    def get_urls(self):
        """Add custom URLs for bulk import"""
        urls = super().get_urls()
        custom_urls = [
            path('bulk-import/', self.admin_site.admin_view(self.bulk_import_view), name='employees_employee_bulk_import'),
            path('download-template/', self.admin_site.admin_view(self.download_template), name='employees_employee_download_template'),
        ]
        return custom_urls + urls

    def bulk_import_employees(self, request, queryset):
        """Admin action to redirect to bulk import page"""
        return redirect('admin:employees_employee_bulk_import')
    bulk_import_employees.short_description = "Bulk Import Employees from File"





    def bulk_import_view(self, request):
        """Handle bulk import of employees"""
        if request.method == 'POST':
            form = BulkImportForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    file = form.cleaned_data['file']
                    file_extension = file.name.split('.')[-1].lower()

                    print(f"Processing file: {file.name}, extension: {file_extension}, size: {file.size}")

                    if file_extension in ['xlsx', 'xls']:
                        result = self._process_excel_file(file)
                    elif file_extension == 'csv':
                        result = self._process_csv_file(file)
                    else:
                        messages.error(request, 'Unsupported file format. Please upload Excel (.xlsx, .xls) or CSV file.')
                        return render(request, 'admin/employees/employee/bulk_import.html', {'form': form})

                    success_count, errors = result
                    print(f"Import result: {success_count} successful, {len(errors)} errors")

                    if success_count > 0:
                        messages.success(request, f'Successfully imported {success_count} employees!')

                    if errors:
                        for error in errors[:10]:  # Show first 10 errors
                            messages.error(request, error)
                        if len(errors) > 10:
                            messages.warning(request, f'... and {len(errors) - 10} more errors.')

                    return redirect('admin:employees_employee_changelist')

                except Exception as e:
                    messages.error(request, f'Error processing file: {str(e)}')
        else:
            form = BulkImportForm()

        return render(request, 'admin/employees/employee/bulk_import.html', {
            'form': form,
            'title': 'Bulk Import Employees',
            'opts': self.model._meta,
        })

    def _process_excel_file(self, file):
        """Process Excel file and create employees"""
        success_count = 0
        errors = []

        try:
            # Read Excel file
            df = pd.read_excel(file)

            # Process each row
            for index, row in df.iterrows():
                try:
                    employee_data = self._extract_employee_data(row, index + 2)  # +2 for header and 0-based index
                    if employee_data:
                        employee = Employee.objects.create(**employee_data)
                        success_count += 1
                except Exception as e:
                    errors.append(f'Row {index + 2}: {str(e)}')

        except Exception as e:
            errors.append(f'Error reading Excel file: {str(e)}')

        return success_count, errors

    def _process_csv_file(self, file):
        """Process CSV file and create employees with multiple encoding support"""
        success_count = 0
        errors = []

        try:
            # Reset file pointer to beginning
            file.seek(0)

            # First, try to detect encoding automatically
            detected_encoding = detect_file_encoding(file)

            # Prepare list of encodings to try
            encodings_to_try = []

            # Add detected encoding first if available
            if detected_encoding:
                encodings_to_try.append(detected_encoding)

            # Add common encodings
            common_encodings = [
                'utf-8',           # Standard UTF-8
                'utf-8-sig',       # UTF-8 with BOM
                'cp1252',          # Windows-1252 (Western European) - most common for Windows
                'latin-1',         # ISO-8859-1 (Western European)
                'iso-8859-1',      # ISO-8859-1
                'cp850',           # DOS Latin-1
                'ascii',           # ASCII
                'utf-16',          # UTF-16
                'utf-32',          # UTF-32
            ]

            # Add common encodings, avoiding duplicates
            for enc in common_encodings:
                if enc not in encodings_to_try:
                    encodings_to_try.append(enc)

            df = None
            encoding_used = None

            for encoding in encodings_to_try:
                try:
                    file.seek(0)
                    df = pd.read_csv(file, encoding=encoding)
                    encoding_used = encoding
                    print(f"Successfully read CSV with encoding: {encoding}")
                    break
                except (UnicodeDecodeError, UnicodeError) as e:
                    print(f"Failed to read with encoding {encoding}: {e}")
                    continue
                except Exception as e:
                    print(f"Error with encoding {encoding}: {e}")
                    continue

            if df is None:
                # If all encodings fail, try with error handling strategies
                fallback_strategies = [
                    ('utf-8', 'replace'),
                    ('latin-1', 'replace'),
                    ('cp1252', 'replace'),
                    ('utf-8', 'ignore'),
                    ('latin-1', 'ignore'),
                ]

                for encoding, error_strategy in fallback_strategies:
                    try:
                        file.seek(0)
                        df = pd.read_csv(file, encoding=encoding, errors=error_strategy)
                        encoding_used = f'{encoding} (with {error_strategy} error handling)'
                        print(f"Read CSV with fallback strategy: {encoding_used}")
                        break
                    except Exception as e:
                        print(f"Fallback strategy {encoding}/{error_strategy} failed: {e}")
                        continue

                if df is None:
                    # Last resort: try to read as binary and convert
                    try:
                        file.seek(0)
                        raw_content = file.read()

                        # Try to decode with different encodings
                        for encoding in ['cp1252', 'latin-1', 'utf-8']:
                            try:
                                decoded_content = raw_content.decode(encoding, errors='replace')
                                # Create a StringIO object for pandas
                                from io import StringIO
                                string_file = StringIO(decoded_content)
                                df = pd.read_csv(string_file)
                                encoding_used = f'{encoding} (binary decode with replace)'
                                print(f"Read CSV with binary decode: {encoding_used}")
                                break
                            except Exception as e:
                                continue
                    except Exception as e:
                        pass

                if df is None:
                    errors.append(
                        'Could not read CSV file with any encoding method. '
                        'Please save your CSV file with UTF-8 encoding and try again. '
                        'In Excel: File → Save As → CSV UTF-8 (Comma delimited)'
                    )
                    return success_count, errors

            # Check if DataFrame is empty
            if df.empty:
                errors.append('CSV file is empty or has no data rows')
                return success_count, errors

            # Log file info for debugging
            print(f"CSV encoding used: {encoding_used}")
            print(f"CSV shape: {df.shape}")
            print(f"CSV columns: {list(df.columns)}")

            # Clean column names (remove BOM, extra spaces, etc.)
            df.columns = df.columns.str.strip()
            df.columns = df.columns.str.replace('\ufeff', '')  # Remove BOM
            print(f"Cleaned CSV columns: {list(df.columns)}")

            # Process each row
            for index, row in df.iterrows():
                try:
                    employee_data = self._extract_employee_data(row, index + 2)  # +2 for header and 0-based index
                    if employee_data:
                        employee = Employee.objects.create(**employee_data)
                        success_count += 1
                        print(f"Successfully created employee: {employee.full_name}")
                except Exception as e:
                    error_msg = f'Row {index + 2}: {str(e)}'
                    errors.append(error_msg)
                    print(f"Error processing row {index + 2}: {e}")

        except Exception as e:
            error_msg = f'Error reading CSV file: {str(e)}'
            errors.append(error_msg)
            print(f"CSV processing error: {e}")

        return success_count, errors

    def _extract_employee_data(self, row, row_number):
        """Extract and validate employee data from a row"""
        try:
            # Create a case-insensitive column mapping
            row_dict = {}
            for key, value in row.items():
                if pd.notna(key):  # Skip NaN keys
                    row_dict[str(key).lower().strip()] = value

            # Define flexible column mappings
            column_mappings = {
                'first_name': ['first_name', 'firstname', 'fname'],
                'middle_name': ['middle_name', 'middlename', 'mname'],
                'last_name': ['last_name', 'lastname', 'lname', 'surname'],
                'email': ['email', 'email_address', 'e_mail'],
                'phone_number': ['phone_number', 'phone', 'mobile', 'telephone'],
                'national_id': ['national_id', 'id_number', 'nationalid', 'id'],
                'address': ['address', 'physical_address', 'location'],
                'gender': ['gender', 'sex'],
                'marital_status': ['marital_status', 'marital', 'marriage_status'],
                'date_of_birth': ['date_of_birth', 'dob', 'birth_date', 'birthdate'],
                'department': ['department', 'department_name', 'dept'],
                'job_title': ['job_title', 'position_title', 'position', 'title', 'role'],
                'employment_type': ['employment_type', 'emp_type', 'type'],
                'date_hired': ['date_hired', 'hire_date', 'start_date', 'employment_date'],
                'kra_pin': ['kra_pin', 'kra', 'pin'],
                'nssf_number': ['nssf_number', 'nssf', 'nssf_no'],
                'shif_number': ['shif_number', 'nhif_number', 'shif', 'nhif'],
                'bank_name': ['bank_name', 'bank'],
                'bank_branch': ['bank_branch', 'branch'],
                'account_number': ['account_number', 'account_no', 'account']
            }

            def get_field_value(field_name):
                """Get field value using flexible column mapping"""
                possible_columns = column_mappings.get(field_name, [field_name])
                for col in possible_columns:
                    if col in row_dict:
                        return row_dict[col]
                return None

            # Required fields with flexible mapping
            first_name = clean_string_value(get_field_value('first_name'))
            last_name = clean_string_value(get_field_value('last_name'))

            if not first_name or not last_name:
                raise ValueError('First name and last name are required')

            # Get or create department and job title
            department_name = clean_string_value(get_field_value('department'))
            job_title_name = clean_string_value(get_field_value('job_title'))

            if not department_name:
                raise ValueError('Department is required')
            if not job_title_name:
                raise ValueError('Job title is required')

            # Get first active organization (or create default)
            organization = Organization.objects.filter(is_active=True).first()
            if not organization:
                raise ValueError('No active organization found. Please create an organization first.')

            # Get or create department
            department, created = Department.objects.get_or_create(
                name=department_name,
                organization=organization,
                defaults={'code': department_name[:10].upper()}
            )

            # Get or create job title
            job_title, created = JobTitle.objects.get_or_create(
                title=job_title_name
            )

            # Prepare employee data
            employee_data = {
                'first_name': first_name,
                'last_name': last_name,
                'department': department,
                'job_title': job_title,
                'employment_type': clean_string_value(get_field_value('employment_type')) or 'PERMANENT',
                'is_active': True,
            }

            # Normalize employment type
            if employee_data['employment_type']:
                emp_type = employee_data['employment_type'].upper()
                if emp_type in ['CONTRACT', 'PERMANENT', 'CASUAL', 'INTERN']:
                    employee_data['employment_type'] = emp_type
                else:
                    employee_data['employment_type'] = 'PERMANENT'

            # Optional fields with flexible handling
            middle_name = clean_string_value(get_field_value('middle_name'))
            if middle_name:
                employee_data['middle_name'] = middle_name

            email = clean_string_value(get_field_value('email'))
            if email and '@' in email:
                employee_data['email'] = email

            phone_number = clean_string_value(get_field_value('phone_number'))
            if phone_number:
                # Clean phone number - remove spaces, dashes, etc.
                phone_clean = re.sub(r'[^\d+]', '', phone_number)
                if phone_clean:
                    employee_data['phone_number'] = phone_clean

            national_id = clean_string_value(get_field_value('national_id'))
            if national_id:
                # Handle numeric national IDs that might be in scientific notation
                national_id_clean = str(parse_number_value(national_id) or national_id)
                # Check if employee with this National ID already exists
                if Employee.objects.filter(national_id=national_id_clean).exists():
                    raise ValueError(f'Employee with National ID {national_id_clean} already exists')
                employee_data['national_id'] = national_id_clean

            address = clean_string_value(get_field_value('address'))
            if address:
                employee_data['address'] = address

            gender = clean_string_value(get_field_value('gender'))
            if gender:
                gender_upper = gender.upper()
                if gender_upper in ['M', 'MALE']:
                    employee_data['gender'] = 'M'
                elif gender_upper in ['F', 'FEMALE']:
                    employee_data['gender'] = 'F'
                elif gender_upper in ['O', 'OTHER']:
                    employee_data['gender'] = 'O'

            marital_status = clean_string_value(get_field_value('marital_status'))
            if marital_status:
                marital_upper = marital_status.upper()
                if marital_upper in ['SINGLE', 'MARRIED', 'DIVORCED', 'WIDOWED']:
                    employee_data['marital_status'] = marital_upper

            # Date fields with flexible parsing
            def parse_date_field(field_name):
                """Parse date field with multiple format support"""
                date_value = get_field_value(field_name)
                if pd.notna(date_value) and date_value:
                    date_str = clean_string_value(date_value)
                    if date_str:
                        # Try multiple date formats
                        date_formats = [
                            '%Y-%m-%d',     # 2024-01-15
                            '%d/%m/%Y',     # 15/01/2024
                            '%m/%d/%Y',     # 01/15/2024
                            '%d-%m-%Y',     # 15-01-2024
                            '%Y/%m/%d',     # 2024/01/15
                            '%d.%m.%Y',     # 15.01.2024
                        ]

                        for fmt in date_formats:
                            try:
                                return datetime.strptime(date_str, fmt).date()
                            except ValueError:
                                continue

                        # If it's already a date object
                        if hasattr(date_value, 'date'):
                            return date_value.date()
                        elif hasattr(date_value, 'year'):
                            return date_value
                return None

            date_of_birth = parse_date_field('date_of_birth')
            if date_of_birth:
                employee_data['date_of_birth'] = date_of_birth

            date_hired = parse_date_field('date_hired')
            if date_hired:
                employee_data['date_hired'] = date_hired

            # Statutory information with flexible handling
            kra_pin = clean_string_value(get_field_value('kra_pin'))
            if kra_pin:
                employee_data['kra_pin'] = kra_pin

            nssf_number = clean_string_value(get_field_value('nssf_number'))
            if nssf_number:
                # Handle numeric NSSF numbers
                nssf_clean = str(parse_number_value(nssf_number) or nssf_number)
                employee_data['nssf_number'] = nssf_clean

            shif_number = clean_string_value(get_field_value('shif_number'))
            if shif_number:
                # Handle numeric SHIF/NHIF numbers
                shif_clean = str(parse_number_value(shif_number) or shif_number)
                employee_data['shif_number'] = shif_clean

            # Banking information with number parsing
            bank_name = clean_string_value(get_field_value('bank_name'))
            if bank_name:
                employee_data['bank_name'] = bank_name

            bank_branch = clean_string_value(get_field_value('bank_branch'))
            if bank_branch:
                employee_data['bank_branch'] = bank_branch

            account_number = get_field_value('account_number')
            if account_number is not None:
                # Handle various account number formats
                account_clean = str(parse_number_value(account_number) or clean_string_value(account_number) or '')
                if account_clean and account_clean != 'None':
                    employee_data['account_number'] = account_clean

            # Handle salary fields and store in Employee model
            salary_fields = ['basic_salary', 'salary', 'wage', 'pay']
            for salary_field in salary_fields:
                salary_value = get_field_value(salary_field)
                if salary_value is not None:
                    try:
                        parsed_salary = parse_number_value(salary_value)
                        if parsed_salary is not None:
                            # Store the salary in the employee data
                            employee_data['basic_salary'] = parsed_salary
                            print(f"Row {row_number}: Parsed and storing {salary_field} = {parsed_salary}")
                            break  # Use the first salary field found
                    except Exception as e:
                        print(f"Row {row_number}: Could not parse {salary_field} '{salary_value}': {e}")

            return employee_data

        except Exception as e:
            raise ValueError(f'Error processing row data: {str(e)}')

    def download_template(self, request):
        """Download Excel template for bulk import"""
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employee_Template"

        # Headers
        headers = [
            'first_name', 'middle_name', 'last_name', 'email', 'phone_number',
            'national_id', 'address', 'gender', 'marital_status', 'date_of_birth',
            'department', 'job_title', 'employment_type', 'date_hired',
            'kra_pin', 'nssf_number', 'shif_number',
            'bank_name', 'bank_branch', 'account_number'
        ]

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Add sample data with various number formats
        sample_data = [
            'John', 'Doe', 'Smith', 'john.smith@example.com', '+254712345678',
            '12345678', '123 Main St, Nairobi', 'M', 'SINGLE', '1990-01-15',
            'IT Department', 'Software Developer', 'PERMANENT', '2024-01-01',
            'A123456789B', '1234567890', 'SHIF123456',
            'KCB Bank', 'Nairobi Branch', '1234567890'
        ]

        # Add second sample with quoted and comma-separated numbers
        sample_data_2 = [
            'Jane', 'Mary', 'Doe', 'jane.doe@example.com', '254-723-456-789',
            '"87654321"', '456 Oak Ave, Mombasa', 'F', 'MARRIED', '15/03/1985',
            'HR Department', 'HR Manager', 'CONTRACT', '01/06/2023',
            '"B987654321C"', '"9876543210"', '"SHIF987654"',
            '"Equity Bank"', '"Mombasa Branch"', '"9,876,543,210"'
        ]

        for col, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col, value=value)

        for col, value in enumerate(sample_data_2, 1):
            ws.cell(row=3, column=col, value=value)

        # Add instructions sheet
        instructions_ws = wb.create_sheet("Instructions")
        instructions = [
            "EMPLOYEE BULK IMPORT TEMPLATE INSTRUCTIONS",
            "",
            "REQUIRED FIELDS (must be provided):",
            "- first_name: Employee's first name",
            "- last_name: Employee's last name",
            "- department: Department name (will be created if doesn't exist)",
            "- job_title: Job title (will be created if doesn't exist)",
            "",
            "OPTIONAL FIELDS:",
            "- middle_name: Employee's middle name",
            "- email: Valid email address",
            "- phone_number: Phone number (e.g., +254712345678)",
            "- national_id: National ID number",
            "- address: Physical address",
            "- gender: M (Male), F (Female), O (Other)",
            "- marital_status: SINGLE, MARRIED, DIVORCED, WIDOWED",
            "- date_of_birth: Format: YYYY-MM-DD or DD/MM/YYYY",
            "- employment_type: PERMANENT, CONTRACT, CASUAL, INTERN",
            "- date_hired: Format: YYYY-MM-DD or DD/MM/YYYY",
            "- kra_pin: KRA PIN (format: A123456789B)",
            "- nssf_number: NSSF number",
            "- shif_number: SHIF number",
            "- bank_name: Bank name",
            "- bank_branch: Bank branch",
            "- account_number: Bank account number",
            "",
            "NOTES:",
            "- Empty cells for optional fields will be ignored",
            "- Duplicate National IDs will be rejected",
            "- Departments and job titles will be created automatically if they don't exist",
            "",
            "NUMBER FORMAT SUPPORT:",
            "- Quoted numbers: \"7000\", \"7,000\", \"7.50\"",
            "- Unquoted numbers: 7000, 7,000, 7.50",
            "- Scientific notation: 1.23E+12",
            "- Phone numbers: +254712345678, 254-712-345-678",
            "- Account numbers: 1234567890, \"1,234,567,890\"",
            "",
            "FLEXIBLE COLUMN NAMES:",
            "- first_name, firstname, fname",
            "- phone_number, phone, mobile",
            "- national_id, id_number, id",
            "- department, department_name, dept",
            "- job_title, position_title, position",
            "- date_hired, hire_date, start_date",
            "- And many more variations..."
        ]

        for row, instruction in enumerate(instructions, 1):
            instructions_ws.cell(row=row, column=1, value=instruction)

        # Adjust column widths
        for ws_sheet in [ws, instructions_ws]:
            for column in ws_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_sheet.column_dimensions[column_letter].width = adjusted_width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="employee_bulk_import_template.xlsx"'

        # Save workbook to response
        wb.save(response)
        return response


@admin.register(SalaryStructure)
class SalaryStructureAdmin(admin.ModelAdmin):
    list_display = [
        'employee', 'basic_salary', 'gross_salary', 'effective_date', 'is_active'
    ]
    list_filter = ['is_active', 'effective_date']
    search_fields = ['employee__first_name', 'employee__last_name', 'employee__employee_number']
    ordering = ['-effective_date']

    fieldsets = (
        ('Employee', {
            'fields': ('employee', 'effective_date', 'is_active')
        }),
        ('Basic Salary', {
            'fields': ('basic_salary',)
        }),
        ('Allowances', {
            'fields': (
                'house_allowance', 'transport_allowance', 'medical_allowance',
                'lunch_allowance', 'communication_allowance', 'other_allowances'
            )
        }),
        ('Benefits (Non-Cash)', {
            'fields': ('car_benefit_value', 'housing_benefit_value')
        }),
        ('Insurance & Pension', {
            'fields': (
                'life_insurance_premium', 'health_insurance_premium',
                'education_insurance_premium', 'pension_contribution'
            )
        }),
        ('Tax Relief Items', {
            'fields': ('mortgage_interest', 'post_retirement_medical_fund')
        }),
    )
