from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse
from django.template.loader import render_to_string
from .models import Employee, Department, JobTitle, SalaryStructure
from .forms import EmployeeForm, SalaryStructureForm, DepartmentForm, JobTitleForm, EmployeeSearchForm, BulkEmployeeImportForm


@login_required
def employee_list(request):
    """List all employees with search and filtering - Login required"""
    form = EmployeeSearchForm(request.GET)
    employees = Employee.objects.select_related('department', 'job_title').all()

    if form.is_valid():
        search = form.cleaned_data.get('search')
        department = form.cleaned_data.get('department')
        employment_type = form.cleaned_data.get('employment_type')
        is_active = form.cleaned_data.get('is_active')

        if search:
            employees = employees.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(payroll_number__icontains=search) |
                Q(email__icontains=search) |
                Q(kra_pin__icontains=search)
            )

        if department:
            employees = employees.filter(department=department)

        if employment_type:
            employees = employees.filter(employment_type=employment_type)

        if is_active:
            employees = employees.filter(is_active=(is_active == 'true'))

    # Pagination
    paginator = Paginator(employees, 25)  # Show 25 employees per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get organization data
    from .models import Organization
    organization = Organization.objects.filter(is_active=True).first()

    context = {
        'page_obj': page_obj,
        'form': form,
        'total_employees': employees.count(),
        'organization': organization,
    }
    return render(request, 'employees/employee_list.html', context)


@login_required
def employee_detail(request, pk):
    """Display employee details - Login required"""
    employee = get_object_or_404(Employee, pk=pk)
    salary_structure = getattr(employee, 'salary_structure', None)

    context = {
        'employee': employee,
        'salary_structure': salary_structure
    }
    return render(request, 'employees/employee_detail.html', context)


@staff_member_required
def employee_create(request):
    """Create a new employee - Admin only"""
    # Ensure we have departments and job titles
    from .models import Department, JobTitle, Organization

    # Get or create default organization
    organization, created = Organization.objects.get_or_create(
        name="Garissa County Government",
        defaults={
            'organization_type': 'government',
            'address_line_1': 'Garissa Town',
            'city': 'Garissa',
            'country': 'Kenya',
            'phone_number': '+254-123-456789',
            'email': 'info@garissa.go.ke',
            'is_active': True
        }
    )

    # Create default departments if none exist
    if not Department.objects.filter(organization=organization).exists():
        default_departments = [
            {'name': 'Administration', 'code': 'ADMIN', 'description': 'Administrative Department'},
            {'name': 'Finance', 'code': 'FIN', 'description': 'Finance and Accounting Department'},
            {'name': 'Human Resources', 'code': 'HR', 'description': 'Human Resources Department'},
            {'name': 'ICT', 'code': 'ICT', 'description': 'Information and Communication Technology'},
            {'name': 'Health Services', 'code': 'HEALTH', 'description': 'Health Services Department'},
            {'name': 'Education', 'code': 'EDU', 'description': 'Education Department'},
            {'name': 'Agriculture', 'code': 'AGRI', 'description': 'Agriculture and Livestock Department'},
            {'name': 'Water', 'code': 'WATER', 'description': 'Water and Sanitation Department'},
            {'name': 'Roads', 'code': 'ROADS', 'description': 'Roads and Infrastructure Department'},
            {'name': 'Ugatuzi', 'code': 'UGAT', 'description': 'Ugatuzi Department'},
            {'name': 'Municipality', 'code': 'MUNIC', 'description': 'Municipality Department'},
        ]

        for dept_data in default_departments:
            Department.objects.create(organization=organization, **dept_data)

    # Create default job titles if none exist
    if not JobTitle.objects.exists():
        default_job_titles = [
            {'title': 'County Secretary', 'description': 'Chief Administrative Officer'},
            {'title': 'Chief Officer', 'description': 'Department Head'},
            {'title': 'Director', 'description': 'Department Director'},
            {'title': 'Manager', 'description': 'Department Manager'},
            {'title': 'Officer', 'description': 'General Officer'},
            {'title': 'Assistant', 'description': 'Assistant Officer'},
            {'title': 'Clerk', 'description': 'Administrative Clerk'},
            {'title': 'Driver', 'description': 'Government Driver'},
            {'title': 'Security Guard', 'description': 'Security Personnel'},
            {'title': 'Casual Worker', 'description': 'Casual/Temporary Worker'},
        ]

        for job_data in default_job_titles:
            JobTitle.objects.create(**job_data)

    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            employee = form.save()
            messages.success(request, f'Employee {employee.full_name} created successfully!')
            return redirect('employees:employee_detail', pk=employee.pk)
    else:
        form = EmployeeForm()

    context = {
        'form': form,
        'title': 'Add New Employee',
        'departments_count': Department.objects.count(),
        'job_titles_count': JobTitle.objects.count(),
        'organization': organization,
    }
    return render(request, 'employees/employee_form.html', context)


@staff_member_required
def employee_update(request, pk):
    """Update an existing employee - Admin only"""
    employee = get_object_or_404(Employee, pk=pk)

    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            employee = form.save()
            messages.success(request, f'Employee {employee.full_name} updated successfully!')
            return redirect('employees:employee_detail', pk=employee.pk)
    else:
        form = EmployeeForm(instance=employee)

    context = {
        'form': form,
        'employee': employee,
        'title': f'Edit {employee.full_name}'
    }
    return render(request, 'employees/employee_form.html', context)


@staff_member_required
def employee_deactivate(request, pk):
    """Deactivate an employee - Admin only"""
    employee = get_object_or_404(Employee, pk=pk)

    if request.method == 'POST':
        employee.is_active = False
        employee.save()
        messages.success(request, f'Employee {employee.full_name} has been deactivated.')
        return redirect('employee_list')

    context = {
        'employee': employee,
        'title': f'Deactivate {employee.full_name}'
    }
    return render(request, 'employees/employee_confirm_deactivate.html', context)


@staff_member_required
def salary_structure_create(request, employee_pk):
    """Create salary structure for an employee - Admin only"""
    employee = get_object_or_404(Employee, pk=employee_pk)

    if request.method == 'POST':
        form = SalaryStructureForm(request.POST)
        if form.is_valid():
            salary_structure = form.save()
            messages.success(request, f'Salary structure created for {employee.full_name}!')
            return redirect('employees:employee_detail', pk=employee.pk)
    else:
        form = SalaryStructureForm(initial={'employee': employee})

    context = {
        'form': form,
        'employee': employee,
        'title': f'Create Salary Structure for {employee.full_name}'
    }
    return render(request, 'employees/salary_structure_form.html', context)


@staff_member_required
def salary_structure_update(request, pk):
    """Update salary structure - Admin only"""
    salary_structure = get_object_or_404(SalaryStructure, pk=pk)

    if request.method == 'POST':
        form = SalaryStructureForm(request.POST, instance=salary_structure)
        if form.is_valid():
            salary_structure = form.save()
            messages.success(request, f'Salary structure updated for {salary_structure.employee.full_name}!')
            return redirect('employees:employee_detail', pk=salary_structure.employee.pk)
    else:
        form = SalaryStructureForm(instance=salary_structure)

    context = {
        'form': form,
        'salary_structure': salary_structure,
        'employee': salary_structure.employee,
        'title': f'Edit Salary Structure for {salary_structure.employee.full_name}'
    }
    return render(request, 'employees/salary_structure_form.html', context)


@staff_member_required
def bulk_employee_import(request):
    """Bulk import employees from Excel file - Admin only"""
    if request.method == 'POST':
        form = BulkEmployeeImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                employees_data, errors = form.process_excel_file()

                if errors:
                    # Show errors to user
                    for error in errors:
                        messages.error(request, error)

                    context = {
                        'form': form,
                        'title': 'Bulk Import Employees',
                        'errors': errors,
                        'processed_count': len(employees_data)
                    }
                    return render(request, 'employees/bulk_import.html', context)

                # Create employees
                created_count = 0
                creation_errors = []

                for emp_data in employees_data:
                    try:
                        # Check if employee with this National ID already exists
                        if Employee.objects.filter(national_id=emp_data['national_id']).exists():
                            creation_errors.append(f'Employee with National ID {emp_data["national_id"]} already exists')
                            continue

                        # Create employee
                        employee = Employee.objects.create(**emp_data)
                        created_count += 1

                    except Exception as e:
                        creation_errors.append(f'Error creating employee {emp_data["first_name"]} {emp_data["last_name"]}: {str(e)}')

                # Show results
                if created_count > 0:
                    messages.success(request, f'Successfully imported {created_count} employees!')

                if creation_errors:
                    for error in creation_errors:
                        messages.warning(request, error)

                return redirect('employees:employee_list')

            except ValidationError as e:
                messages.error(request, str(e))
    else:
        form = BulkEmployeeImportForm()

    # Get organization for logo display
    from .models import Organization
    try:
        organization = Organization.objects.filter(is_active=True).first()
    except:
        organization = None

    context = {
        'form': form,
        'title': 'Bulk Import Employees',
        'organization': organization,
    }
    return render(request, 'employees/bulk_import.html', context)


@staff_member_required
def download_employee_template(request):
    """Download Excel template for bulk employee import"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO

    # Create workbook
    wb = Workbook()

    # Employee Template Sheet
    ws = wb.active
    ws.title = "Employee_Template"

    # Headers
    headers = [
        'first_name', 'middle_name', 'last_name', 'national_id', 'phone_number',
        'email', 'gender', 'date_of_birth', 'address', 'department', 'job_title',
        'employment_type', 'date_hired', 'kra_pin', 'nssf_number', 'shif_number',
        'bank_code', 'bank_name', 'bank_branch', 'account_number'
    ]

    # Add headers with formatting
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Sample data
    sample_data = [
        ['John', 'Doe', 'Mwangi', '12345678', '+254712345678', 'john.mwangi@garissa.go.ke',
         'M', '1990-01-15', 'P.O. Box 123, Garissa', 'Administration', 'Officer',
         'permanent', '2023-01-01', 'A123456789B', '123456789', 'SHIF123456',
         '68058', 'Equity Bank', 'Garissa Branch', '1234567890'],
        ['Jane', 'Smith', 'Wanjiku', '87654321', '+254787654321', 'jane.wanjiku@garissa.go.ke',
         'F', '1985-05-20', 'P.O. Box 456, Garissa', 'Finance', 'Manager',
         'contract', '2022-06-15', 'C987654321D', '987654321', 'SHIF789012',
         '01169', 'KCB Bank', 'Garissa Branch', '0987654321']
    ]

    # Add sample data
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Instructions Sheet
    instructions_ws = wb.create_sheet("Instructions")
    instructions = [
        '1. Fill in employee data in the Employee_Template sheet',
        '2. MANDATORY FIELDS (Required): first_name, last_name, national_id, department, job_title, employment_type, bank_name, bank_branch, account_number',
        '3. OPTIONAL FIELDS: middle_name, phone_number, email, gender, date_of_birth, address, date_hired, kra_pin, nssf_number, shif_number, bank_code',
        '4. UNIQUE FIELDS (Must be unique across all employees):',
        '   - National ID (MANDATORY & UNIQUE): 8 digits, cannot be duplicated',
        '   - Bank Account Number (MANDATORY & UNIQUE): Cannot be duplicated',
        '   - KRA PIN (OPTIONAL & UNIQUE): If provided, cannot be duplicated',
        '   - SHIF Number (OPTIONAL & UNIQUE): If provided, cannot be duplicated',
        '   - NSSF Number (OPTIONAL & UNIQUE): If provided, cannot be duplicated',
        '5. Department must match exactly: Administration, Finance, Human Resources, ICT, Health Services, Education, Agriculture, Water, Roads, Ugatuzi, Municipality',
        '6. Job Title must match exactly: County Secretary, Chief Officer, Director, Manager, Officer, Assistant, Clerk, Driver, Security Guard, Casual Worker',
        '7. Employment Type: permanent, contract, casual, intern',
        '8. Gender: M or F (optional)',
        '9. Date format: YYYY-MM-DD (e.g., 2023-01-15)',
        '10. National ID: Exactly 8 digits (e.g., 12345678)',
        '11. KRA PIN format: A123456789B (optional but unique if provided)',
        '12. Bank Code (optional): 12053 (National Bank), 68058 (Equity Bank), 01169 (KCB Bank), 11081 (Cooperative Bank), 03017 (Absa Bank), 74004 (Premier Bank), 72006 (Gulf African Bank)',
        '13. Phone format: +254712345678',
        '14. Remove sample data before uploading your actual employee data',
        '15. Ensure all unique fields have different values for each employee'
    ]

    # Add instructions to sheet
    for row_idx, instruction in enumerate(instructions, 1):
        instructions_ws.cell(row=row_idx, column=1, value=instruction)

    # Departments reference sheet
    dept_ws = wb.create_sheet("Departments")
    dept_ws.cell(row=1, column=1, value="Available_Departments").font = Font(bold=True)
    departments = [
        'Administration', 'Finance', 'Human Resources', 'ICT',
        'Health Services', 'Education', 'Agriculture', 'Water',
        'Roads', 'Ugatuzi', 'Municipality'
    ]
    for row_idx, dept in enumerate(departments, 2):
        dept_ws.cell(row=row_idx, column=1, value=dept)

    # Job titles reference sheet
    job_ws = wb.create_sheet("Job_Titles")
    job_ws.cell(row=1, column=1, value="Available_Job_Titles").font = Font(bold=True)
    job_titles = [
        'County Secretary', 'Chief Officer', 'Director', 'Manager',
        'Officer', 'Assistant', 'Clerk', 'Driver',
        'Security Guard', 'Casual Worker'
    ]
    for row_idx, title in enumerate(job_titles, 2):
        job_ws.cell(row=row_idx, column=1, value=title)

    # Bank codes reference sheet
    bank_ws = wb.create_sheet("Bank_Codes")
    bank_ws.cell(row=1, column=1, value="Bank_Code").font = Font(bold=True)
    bank_ws.cell(row=1, column=2, value="Bank_Name").font = Font(bold=True)

    bank_data = [
        ('12053', 'National Bank'), ('68058', 'Equity Bank'), ('01169', 'KCB Bank'),
        ('11081', 'Cooperative Bank'), ('03017', 'Absa Bank'), ('74004', 'Premier Bank'),
        ('72006', 'Gulf African Bank')
    ]
    for row_idx, (code, name) in enumerate(bank_data, 2):
        bank_ws.cell(row=row_idx, column=1, value=code)
        bank_ws.cell(row=row_idx, column=2, value=name)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="employee_import_template.xlsx"'

    return response


@staff_member_required
def employee_delete(request, pk):
    """Delete a single employee - Admin only"""
    employee = get_object_or_404(Employee, pk=pk)

    if request.method == 'POST':
        employee_name = employee.full_name
        employee.delete()
        messages.success(request, f'Employee {employee_name} has been deleted successfully.')
        return redirect('employees:employee_list')

    context = {
        'employee': employee,
        'title': f'Delete Employee - {employee.full_name}'
    }
    return render(request, 'employees/employee_delete.html', context)


@staff_member_required
def bulk_employee_delete(request):
    """Bulk delete employees - Admin only"""
    if request.method == 'POST':
        employee_ids = request.POST.getlist('employee_ids')

        if not employee_ids:
            messages.error(request, 'No employees selected for deletion.')
            return redirect('employees:employee_list')

        try:
            # Get employee names for confirmation message
            employees = Employee.objects.filter(id__in=employee_ids)
            employee_names = [emp.full_name for emp in employees]
            count = employees.count()

            # Delete the employees
            employees.delete()

            if count == 1:
                messages.success(request, f'Employee {employee_names[0]} has been deleted successfully.')
            else:
                messages.success(request, f'{count} employees have been deleted successfully.')

        except Exception as e:
            messages.error(request, f'Error deleting employees: {str(e)}')

    return redirect('employees:employee_list')


@login_required
def employee_search_ajax(request):
    """AJAX endpoint for real-time employee search"""
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form = EmployeeSearchForm(request.GET)
        employees = Employee.objects.select_related('department', 'job_title').all()

        if form.is_valid():
            search = form.cleaned_data.get('search')
            department = form.cleaned_data.get('department')
            employment_type = form.cleaned_data.get('employment_type')
            is_active = form.cleaned_data.get('is_active')

            if search:
                employees = employees.filter(
                    Q(first_name__icontains=search) |
                    Q(last_name__icontains=search) |
                    Q(payroll_number__icontains=search) |
                    Q(email__icontains=search) |
                    Q(kra_pin__icontains=search)
                )

            if department:
                employees = employees.filter(department=department)

            if employment_type:
                employees = employees.filter(employment_type=employment_type)

            if is_active:
                employees = employees.filter(is_active=(is_active == 'true'))

        # Pagination for AJAX
        paginator = Paginator(employees, 25)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        # Render the employee table partial
        html = render_to_string('employees/partials/employee_table.html', {
            'page_obj': page_obj,
            'total_employees': employees.count()
        }, request=request)

        return JsonResponse({
            'html': html,
            'total_employees': employees.count(),
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages
        })

    return JsonResponse({'error': 'Invalid request'}, status=400)
